from flask import Flask, render_template, request, redirect, session, send_file
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from openpyxl import load_workbook
import random
import os

app = Flask(__name__)
app.secret_key = "secret123"

app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///database.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
db = SQLAlchemy(app)


# ======================
# DATABASE MODELS
# ======================
class Subject(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True)
    exam_timer = db.Column(db.Integer, default=120)
    start_time = db.Column(db.String(20))
    end_time = db.Column(db.String(20))


class Question(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    subject = db.Column(db.String(100))
    question = db.Column(db.String(500))
    opt1 = db.Column(db.String(200))
    opt2 = db.Column(db.String(200))
    opt3 = db.Column(db.String(200))
    opt4 = db.Column(db.String(200))
    answer = db.Column(db.String(200))


class Result(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100))
    student_id = db.Column(db.String(50))
    score = db.Column(db.Integer)
    total = db.Column(db.Integer)
    subject = db.Column(db.String(100))
    date = db.Column(db.DateTime, default=datetime.utcnow)


with app.app_context():
    db.create_all()

    if Subject.query.count() == 0:
        db.session.add(
            Subject(
                name="CSC 201",
                exam_timer=120,
                start_time="08:00",
                end_time="18:00"
            )
        )
        db.session.commit()


# ======================
# HELPERS
# ======================
def current_time_str():
    return datetime.now().strftime("%H:%M")


def subject_is_open(subject_obj):
    now = current_time_str()
    return subject_obj.start_time <= now <= subject_obj.end_time


# ======================
# LOGIN
# ======================
@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        student_id = request.form.get("student_id", "").strip()
        password = request.form.get("password", "").strip()

        if username.lower() == "admin" and password == "admin123":
            session.clear()
            session["admin"] = True
            return redirect("/admin")

        if username:
            session.clear()
            session["student"] = username
            session["student_id"] = student_id
            return redirect("/home")

    return render_template("login.html")


# ======================
# HOME
# ======================
@app.route("/home")
def home():
    if "student" not in session:
        return redirect("/")

    subjects = Subject.query.all()
    return render_template(
        "home.html",
        student_name=session["student"],
        subjects=subjects,
        now=current_time_str()
    )


# ======================
# STUDENT EXAM
# ======================
@app.route("/student/<int:subject_id>")
def student(subject_id):
    if "student" not in session:
        return redirect("/")

    subject = Subject.query.get_or_404(subject_id)

    if not subject_is_open(subject):
        return render_template(
            "subject_closed.html",
            subject=subject,
            now=current_time_str(),
            student_name=session["student"]
        )

    questions = Question.query.filter_by(subject=subject.name).all()
    random.shuffle(questions)

    return render_template(
        "student.html",
        questions=questions,
        student_name=session["student"],
        exam_timer=subject.exam_timer,
        subject_name=subject.name,
        subject_id=subject.id
    )


# ======================
# REVIEW
# ======================
@app.route("/review/<int:subject_id>", methods=["POST"])
def review(subject_id):
    if "student" not in session:
        return redirect("/")

    subject = Subject.query.get_or_404(subject_id)
    questions = Question.query.filter_by(subject=subject.name).all()
    answers = {str(q.id): request.form.get(str(q.id)) for q in questions}

    session["answers"] = answers
    session["current_subject"] = subject.name
    session["current_subject_id"] = subject.id

    return render_template(
        "review.html",
        questions=questions,
        answers=answers,
        student_name=session["student"],
        subject_name=subject.name,
        subject_id=subject.id
    )


# ======================
# SUBMIT
# ======================
@app.route("/submit", methods=["POST"])
def submit():
    if "student" not in session:
        return redirect("/")

    subject_name = session.get("current_subject")
    if not subject_name:
        return redirect("/home")

    questions = Question.query.filter_by(subject=subject_name).all()
    answers = session.get("answers", {})

    score = 0
    for q in questions:
        selected = answers.get(str(q.id))
        if selected == q.answer:
            score += 1

    result = Result(
        name=session["student"],
        student_id=session.get("student_id", ""),
        score=score,
        total=len(questions),
        subject=subject_name
    )
    db.session.add(result)
    db.session.commit()

    return render_template(
        "result.html",
        score=score,
        total=len(questions),
        student_name=session["student"],
        subject_name=subject_name,
        result_id=result.id
    )


# ======================
# ADMIN
# ======================
@app.route("/admin")
def admin():
    if "admin" not in session:
        return redirect("/")

    questions = Question.query.all()
    subjects = Subject.query.all()

    return render_template(
        "admin.html",
        questions=questions,
        subjects=subjects
    )


# ======================
# RESULTS DASHBOARD
# ======================
@app.route("/results")
def results():
    if "admin" not in session:
        return redirect("/")

    all_results = Result.query.order_by(Result.date.desc()).all()

    total_results = len(all_results)
    total_score_sum = sum(r.score for r in all_results)
    average_score = round(total_score_sum / total_results, 2) if total_results > 0 else 0
    highest_score = max((r.score for r in all_results), default=0)
    lowest_score = min((r.score for r in all_results), default=0)

    return render_template(
        "results.html",
        results=all_results,
        total_results=total_results,
        average_score=average_score,
        highest_score=highest_score,
        lowest_score=lowest_score
    )


# ======================
# LEADERBOARD
# ======================
@app.route("/leaderboard")
def leaderboard():
    results = Result.query.order_by(Result.score.desc(), Result.date.asc()).all()
    return render_template("leaderboard.html", results=results)


# ======================
# PDF RESULT
# ======================
@app.route("/download_result/<int:result_id>")
def download_result(result_id):
    if "student" not in session and "admin" not in session:
        return redirect("/")

    result = Result.query.get_or_404(result_id)

    filename = f"result_{result.id}.pdf"
    filepath = os.path.join("/tmp", filename)

    c = canvas.Canvas(filepath, pagesize=A4)
    width, height = A4

    c.setFont("Helvetica-Bold", 18)
    c.drawString(180, height - 80, "CBT RESULT SLIP")

    c.setFont("Helvetica", 12)
    c.drawString(70, height - 140, f"Name: {result.name}")
    c.drawString(70, height - 170, f"Student ID: {result.student_id}")
    c.drawString(70, height - 200, f"Subject: {result.subject}")
    c.drawString(70, height - 230, f"Score: {result.score}/{result.total}")
    c.drawString(70, height - 260, f"Date: {result.date.strftime('%Y-%m-%d %H:%M:%S')}")

    c.setFont("Helvetica-Bold", 11)
    c.drawString(70, height - 320, "Designed by Ifelodun")
    c.drawString(70, height - 340, "Contact: noplysola@gmail.com")

    c.save()

    return send_file(filepath, as_attachment=True, download_name=filename)


# ======================
# SUBJECTS
# ======================
@app.route("/add_subject", methods=["POST"])
def add_subject():
    if "admin" not in session:
        return redirect("/")

    name = request.form.get("name", "").strip()
    exam_timer = request.form.get("exam_timer", "").strip()
    start_time = request.form.get("start_time", "").strip()
    end_time = request.form.get("end_time", "").strip()

    if name and exam_timer.isdigit() and start_time and end_time:
        existing = Subject.query.filter_by(name=name).first()
        if not existing:
            db.session.add(
                Subject(
                    name=name,
                    exam_timer=int(exam_timer),
                    start_time=start_time,
                    end_time=end_time
                )
            )
            db.session.commit()

    return redirect("/admin")


@app.route("/update_subject/<int:subject_id>", methods=["POST"])
def update_subject(subject_id):
    if "admin" not in session:
        return redirect("/")

    subject = Subject.query.get_or_404(subject_id)

    name = request.form.get("name", "").strip()
    exam_timer = request.form.get("exam_timer", "").strip()
    start_time = request.form.get("start_time", "").strip()
    end_time = request.form.get("end_time", "").strip()

    if name and exam_timer.isdigit() and start_time and end_time:
        old_name = subject.name

        subject.name = name
        subject.exam_timer = int(exam_timer)
        subject.start_time = start_time
        subject.end_time = end_time

        questions = Question.query.filter_by(subject=old_name).all()
        for q in questions:
            q.subject = name

        db.session.commit()

    return redirect("/admin")


@app.route("/delete_subject/<int:subject_id>")
def delete_subject(subject_id):
    if "admin" not in session:
        return redirect("/")

    subject = Subject.query.get_or_404(subject_id)
    Question.query.filter_by(subject=subject.name).delete()
    db.session.delete(subject)
    db.session.commit()

    return redirect("/admin")


# ======================
# QUESTIONS
# ======================
@app.route("/add", methods=["GET", "POST"])
def add():
    if "admin" not in session:
        return redirect("/")

    subjects = Subject.query.all()

    if request.method == "POST":
        subject = request.form.get("subject", "").strip()
        question = request.form.get("question", "").strip()
        opt1 = request.form.get("opt1", "").strip()
        opt2 = request.form.get("opt2", "").strip()
        opt3 = request.form.get("opt3", "").strip()
        opt4 = request.form.get("opt4", "").strip()
        answer = request.form.get("answer", "").strip()

        if subject and question and opt1 and opt2 and opt3 and opt4 and answer:
            q = Question(
                subject=subject,
                question=question,
                opt1=opt1,
                opt2=opt2,
                opt3=opt3,
                opt4=opt4,
                answer=answer
            )
            db.session.add(q)
            db.session.commit()

        return redirect("/admin")

    return render_template("add.html", subjects=subjects)


@app.route("/edit_question/<int:question_id>", methods=["GET", "POST"])
def edit_question(question_id):
    if "admin" not in session:
        return redirect("/")

    q = Question.query.get_or_404(question_id)
    subjects = Subject.query.all()

    if request.method == "POST":
        q.subject = request.form.get("subject", "").strip()
        q.question = request.form.get("question", "").strip()
        q.opt1 = request.form.get("opt1", "").strip()
        q.opt2 = request.form.get("opt2", "").strip()
        q.opt3 = request.form.get("opt3", "").strip()
        q.opt4 = request.form.get("opt4", "").strip()
        q.answer = request.form.get("answer", "").strip()

        db.session.commit()
        return redirect("/admin")

    return render_template("edit_question.html", q=q, subjects=subjects)


@app.route("/import_excel", methods=["GET", "POST"])
def import_excel():
    if "admin" not in session:
        return redirect("/")

    subjects = Subject.query.all()

    if request.method == "POST":
        excel_file = request.files.get("excel_file")
        selected_subject = request.form.get("subject", "").strip()

        if excel_file and selected_subject:
            upload_path = os.path.join("/tmp", excel_file.filename)
            excel_file.save(upload_path)

            wb = load_workbook(upload_path)
            ws = wb.active

            for row in ws.iter_rows(min_row=2, values_only=True):
                question_text = row[0]
                opt1 = row[1]
                opt2 = row[2]
                opt3 = row[3]
                opt4 = row[4]
                answer = row[5]

                if question_text and opt1 and opt2 and opt3 and opt4 and answer:
                    q = Question(
                        subject=selected_subject,
                        question=str(question_text).strip(),
                        opt1=str(opt1).strip(),
                        opt2=str(opt2).strip(),
                        opt3=str(opt3).strip(),
                        opt4=str(opt4).strip(),
                        answer=str(answer).strip()
                    )
                    db.session.add(q)

            db.session.commit()
            return redirect("/admin")

    return render_template("import_excel.html", subjects=subjects)


@app.route("/delete/<int:id>")
def delete(id):
    if "admin" not in session:
        return redirect("/")

    q = Question.query.get_or_404(id)
    db.session.delete(q)
    db.session.commit()
    return redirect("/admin")


# ======================
# LOGOUT
# ======================
@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")


port = int(os.environ.get("PORT", 5000))
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=port)
